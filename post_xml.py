#!/usr/bin/python

import glob
import json
import os
import os.path
import requests
import sys
import time
import xml.dom.minidom
import yaml
import openpyxl


try:
    cfgFile = sys.argv[1]

except Exception as e:
    print str(e)
    sys.exit(0)

with open( cfgFile, 'r' ) as config:
    config = yaml.safe_load( config )


auth = {
    'aaaUser': {
        'attributes': {
            'name': config['name'],
            'pwd': config['passwd']
            }
        }
    }

print config['name']
print config['passwd']
print config['host']

status = 0
while( status != 200 ):
    url = 'http://%s/api/aaaLogin.json' % config['host']
    while(1):
        try:
            r = requests.post( url, data=json.dumps(auth), timeout=1, verify=False )
            break;
        except Exception as e:
            print "timeout"
    status = r.status_code
    print r.text
    cookies = r.cookies
    time.sleep(1)

def runConfig( status, config ):
    tasks = config['tests']
    exlFile = config['exl-wb']

    try:
        print '++++++++ opening Excel WB (%s) ++++++++' % exlFile
        wb = openpyxl.load_workbook(exlFile, data_only=True)
    except Exception as e:
            print "Error Can't Open file %s" % exlFile
    except:
            print "error opening excel %s data input" % exlFile

    for i in range(0,5,1):
        current_sheet = wb.get_sheet_names()[i]
        print 'First sheet name = %s' % current_sheet
        active_sheet = wb.get_sheet_by_name(current_sheet)
        print active_sheet
        active_dict = {}
        active_dict[active_sheet] = []
        header = []
        max_col = active_sheet.max_column
        max_row = active_sheet.max_row
        for col_num in range(1, max_col + 1):
            header.append(active_sheet.cell(row=1, column=col_num).value)
            for row_num in range(2, max_row + 1):
                new_entry = {}
                for col_num in range(1, max_col + 1):
                    if active_sheet.cell(row=row_num, column=col_num).value != None:
                        new_entry[header[col_num - 1]] = active_sheet.cell(row=row_num, column=col_num).value
                    else:
                        new_entry[header[col_num - 1]] = ""
                active_dict[active_sheet].append(new_entry)

    print active_dict

    """
    print sheet title 
    print(sheet.title)
    """
    print '++++++++ Active Sheet (%s) ++++++++' % active_sheet



with open( cfgFile, 'r' ) as payload:
    if( status==200):
        time.sleep(5)
    else:
        raw_input( 'Hit return to process %s' % cfgFile )

        data = payload.read()
        print '++++++++ REQUEST (%s) ++++++++' % cfgFile
        print data
        print '-------- REQUEST (%s) --------' % cfgFile
        url = 'http://%s/api/node/mo/.xml' % config['host']
        r = requests.post( url,
                                   cookies=cookies,
                                   data=data, verify=False )
        result = xml.dom.minidom.parseString( r.text )
        status = r.status_code
        print '++++++++ RESPONSE (%s) ++++++++' % cfgFile
        print result.toprettyxml()
        print '-------- RESPONSE (%s) --------' % cfgFile
        print status

runConfig( status, config )
